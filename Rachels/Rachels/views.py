# views.py
from datetime import datetime, date
import csv
from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Count, Sum
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.core.paginator import Paginator
from .forms import *
from .models import *
from decimal import Decimal
from django.http import JsonResponse
from .models import Notification
from django.contrib.auth import get_user_model
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from django.db.models import Max, Q
from django.contrib import admin
from django.db import transaction
from openpyxl import Workbook
import re

User = get_user_model()

def _normalize_location_for_group(loc):
    if not loc:
        return ""
    return ''.join(ch.lower() if ch.isalnum() else '_' for ch in loc).strip('_')

def user_is_admin(user):
    return user.is_authenticated and user.is_superuser

def user_in_manager_group_for_location(user, location):
    if not user.is_authenticated:
        return False
    grp_name = f"manager_{_normalize_location_for_group(location)}"
    return user.groups.filter(name=grp_name).exists()

def user_can_view_location(user, store):
    # Admin can see everything
    if user.is_superuser:
        return True

    # Manager can only see their own store
    if hasattr(user, "managerprofile"):
        return user.managerprofile.store == store

    return False

def admin_required(view_func):
    return user_passes_test(user_is_admin)(view_func)

@login_required
def home(request):
    user = request.user

    # ----------------------------
    # ROLE / STORE SCOPE
    # ----------------------------
    user_store = None
    if not user.is_superuser and hasattr(user, "managerprofile"):
        user_store = user.managerprofile.store

    # ----------------------------
    # Check if Record has status field (backward-safe)
    # ----------------------------
    field_names = [f.name for f in Record._meta.fields]
    has_status = "status" in field_names

    base_qs = Record.objects.all()

    # 🔒 Restrict managers to their store
    if user_store:
        base_qs = base_qs.filter(location=user_store)

    # ----------------------------
    # TOTAL PENDING
    # ----------------------------
    if has_status:
        total_pending = base_qs.filter(status__iexact="Pending").count()
    else:
        total_pending = base_qs.count()

    # ----------------------------
    # PENDING BY STORE
    # ----------------------------
    if has_status:
        pending_by_location = (
            base_qs
            .filter(status__iexact="Pending")
            .values("location__name")
            .annotate(count=Count("id"))
            .order_by("-count", "location__name")
        )
    else:
        pending_by_location = (
            base_qs
            .values("location__name")
            .annotate(count=Count("id"))
            .order_by("-count", "location__name")
        )

    # ----------------------------
    # TOP 5 PENDING ORDERS
    # ----------------------------
    if has_status:
        top5_orders = (
            base_qs
            .filter(status__iexact="Pending")
            .select_related("location")
            .order_by("-date", "-id")[:5]
        )
    else:
        top5_orders = (
            base_qs
            .select_related("location")
            .order_by("-date", "-id")[:5]
        )

    # ----------------------------
    # LATEST RECORDS
    # ----------------------------
    latest_records = (
        base_qs
        .select_related("location")
        .order_by("-date", "-id")[:5]
    )

    # ----------------------------
    # STORE CARDS
    # ----------------------------
    if user_store:
        stores = Store.objects.filter(id=user_store.id, is_active=True)
    else:
        stores = Store.objects.filter(is_active=True)

    location_cards = []

    for store in stores:
        if has_status:
            pending_base = (
                base_qs
                .filter(location=store, status__iexact="Pending")
                .order_by("-date", "-id")
            )

            successful_base = (
                base_qs
                .filter(location=store)
                .exclude(status__iexact="Pending")
                .order_by("-date", "-id")
            )
        else:
            pending_base = (
                base_qs
                .filter(location=store)
                .order_by("-date", "-id")
            )
            successful_base = Record.objects.none()

        location_cards.append({
            "location": store,
            "pending": list(pending_base[:5]),
            "successful": list(successful_base[:5]),
            "pending_count": pending_base.count(),
            "successful_count": successful_base.count(),
        })

    context = {
        "total_pending": total_pending,
        "pending_by_location": pending_by_location,
        "latest_records": latest_records,
        "top5_orders": top5_orders,
        "location_cards": location_cards,
    }

    return render(request, "home.html", context)

@login_required
def show_all_records(request):
    user = request.user

    # ----------------------------
    # ROLE / STORE SCOPE
    # ----------------------------
    user_store = None
    if not user.is_superuser and hasattr(user, "managerprofile"):
        user_store = user.managerprofile.store

    # ----------------------------
    # BASE QUERYSET
    # ----------------------------
    qs = Record.objects.all().order_by('-date', '-id')

    # 🔒 Restrict managers to their store
    if user_store:
        qs = qs.filter(location=user_store)

    # ----------------------------
    # TEXT SEARCH
    # ----------------------------
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(details__icontains=q)

    # ----------------------------
    # LOCATION FILTER (admins only)
    # ----------------------------
    location = request.GET.get('location', '').strip()
    if location and not user_store:
        qs = qs.filter(location__name__iexact=location)

    # ----------------------------
    # STATUS FILTER
    # ----------------------------
    status = request.GET.get('status', '').strip()
    if status:
        qs = qs.filter(status__iexact=status)

    # ----------------------------
    # MONTH FILTER
    # ----------------------------
    if request.GET.get('month') == 'this':
        today = date.today()
        qs = qs.filter(date__year=today.year, date__month=today.month)

    
    elif request.GET.get('month') == 'today':
        today = date.today()
        qs = qs.filter(date=today)

    # ----------------------------
    # PAGINATION
    # ----------------------------
    per_page = 25
    paginator = Paginator(qs, per_page)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # Compact pagination list
    total_pages = paginator.num_pages
    current = page_obj.number if page_obj else 1
    pagination_items = []
    last_was_ellipsis = False

    for p in range(1, total_pages + 1):
        show = (
            p <= 2 or
            p > total_pages - 2 or
            abs(p - current) <= 1
        )
        if show:
            pagination_items.append(p)
            last_was_ellipsis = False
        else:
            if not last_was_ellipsis:
                pagination_items.append('...')
                last_was_ellipsis = True

    context = {
        'records': page_obj.object_list,
        'page_obj': page_obj,
        'paginator': paginator,
        'pagination_items': pagination_items,
        'request': request,
    }

    return render(request, 'DisplayRecord.html', context)

@login_required
def record_detail(request, pk):
    record = get_object_or_404(Record, pk=pk)

    if not user_can_view_location(request.user, record.location):
        return HttpResponseForbidden("You don't have permission to view this record.")

    # Safe total calculation
    total_amount = 0
    if record.item and record.item.unit_price:
        total_amount = record.quantity * record.item.unit_price

    context = {
        "record": record,
        "total_amount": total_amount,
        "is_admin": request.user.is_superuser,
    }

    return render(request, "record_detail.html", context)

@admin_required
def mark_completed(request, pk):
    record = get_object_or_404(Record, pk=pk)
    if request.method == "POST":
        record.status = "Completed"
        record.save()
        messages.success(request, "Record marked completed.")
        return redirect('show_all_records')
    return redirect('record_detail', pk=pk)

@admin_required
def delete_record(request, pk):
    record = get_object_or_404(Record, pk=pk)
    if request.method == "POST":
        record.delete()
        messages.success(request, "Record deleted.")
        return redirect('show_all_records')
    return render(request, "delete_record.html", {"record": record})

def _parse_date(s):
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None

@login_required
def export_form(request):
    user = request.user
    is_admin = user.is_superuser

    manager_store = None
    if not is_admin and hasattr(user, "managerprofile"):
        manager_store = user.managerprofile.store

    initial = {
        'from_date': request.GET.get('from_date', ''),
        'to_date': request.GET.get('to_date', ''),
        'location': request.GET.get('location', ''),
        'status': request.GET.get('status', ''),
    }

    stores = Store.objects.filter(is_active=True)

    return render(
        request,
        "export_records.html",
        {
            'initial': initial,
            'stores': stores,
            'is_admin': is_admin,          # ✅ REQUIRED
            'manager_store': manager_store # optional but safe
        }
    )

def _safe_sheet_name(name: str) -> str:
    # Excel forbids these characters
    name = re.sub(r'[:\\/?*\[\]]', '', name)
    return name[:31] or "Sheet"

def export_excel(request):
    data = request.GET if request.method == "GET" else request.POST

    from_date = _parse_date(data.get('from_date', '').strip())
    to_date = _parse_date(data.get('to_date', '').strip())
    location_id = data.get('location', '').strip()  # STORE ID
    status = data.get('status', '').strip()

    # ---------- VALIDATION ----------
    if from_date and to_date and from_date > to_date:
        messages.error(request, "From date cannot be after To date.")
        return redirect('export_form')

    # ---------- BASE QUERY ----------
    qs = Record.objects.select_related(
        'vendor', 'location', 'item'
    ).order_by('-date', '-id')

    if from_date:
        qs = qs.filter(date__gte=from_date)
    if to_date:
        qs = qs.filter(date__lte=to_date)
    if location_id:
        qs = qs.filter(location_id=location_id)  # ✅ FK-safe
    if status:
        qs = qs.filter(status__iexact=status)

    # ---------- GROUP DATA ----------
    # Store -> Vendor -> Item -> [records]
    grouped = defaultdict(
        lambda: defaultdict(lambda: defaultdict(list))
    )

    for r in qs:
        store = r.location
        store_name = store.name if store else "Unknown Location"

        vendor_name = r.vendor.name if r.vendor else "Unknown Vendor"
        item_name = r.item.item_name if r.item else "Unknown Item"

        grouped[store][vendor_name][item_name].append(r)

    # ---------- CREATE EXCEL ----------
    wb = Workbook()
    wb.remove(wb.active)

    header = ['Order ID', 'Date', 'Status', 'Quantity']
    header_font = Font(bold=True)

    for store, vendors in grouped.items():
        store_name = store.name if store else "Unknown Location"
        ws = wb.create_sheet(title=_safe_sheet_name(store_name))

        row = 1
        ws.cell(row=row, column=1, value=f"Location: {store_name}")
        ws.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 2

        for vendor, items in vendors.items():
            ws.cell(row=row, column=1, value=f"Vendor: {vendor}")
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1

            for item, records in items.items():
                ws.cell(row=row, column=1, value=f"Item: {item}")
                ws.cell(row=row, column=1).font = Font(italic=True)
                row += 1

                # table header
                for col, h in enumerate(header, start=1):
                    cell = ws.cell(row=row, column=col, value=h)
                    cell.font = header_font
                row += 1

                for r in records:
                    ws.append([
                        r.pk,
                        r.date.isoformat() if r.date else '',
                        r.status,
                        r.quantity,
                    ])
                    row += 1

                row += 1  # space after item

            row += 1  # space after vendor

        # autosize columns
        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = 20

    # ---------- RESPONSE ----------
    fd = from_date.isoformat() if from_date else timezone.localdate().isoformat()
    td = to_date.isoformat() if to_date else timezone.localdate().isoformat()
    filename = f"orders-{fd}-{td}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response

@login_required
def add_vendor(request):
    if request.method == "POST":
        form = VendorForm(request.POST)

        items = request.POST.getlist("items[]")
        units = request.POST.getlist("units[]")
        prices = request.POST.getlist("prices[]")

        if form.is_valid() and items:
            vendor = form.save()

            for name, unit, price in zip(items, units, prices):
                if name.strip() and unit and price:
                    VendorItem.objects.create(
                        vendor=vendor,
                        item_name=name.strip(),
                        unit=unit,
                        unit_price=price
                    )

            messages.success(request, "Vendor and items saved successfully.")
            return redirect("add_vendor")

        messages.error(request, "Please provide vendor name and valid items.")

    else:
        form = VendorForm()

    return render(request, "add_vendor.html", {"form": form})

@admin_required
def advance_list(request):
    qs = AdvanceSalary.objects.all()
    total_given = qs.aggregate(total=Sum("amount"))["total"] or 0
    return render(request, "advance_list.html", {
        "advances": qs,
        "total_given": total_given,
    })

@admin_required
def advance_add(request):
    if request.method == "POST":
        form = AdvanceSalaryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Advance saved.")
            return redirect("advance_list")
    else:
        form = AdvanceSalaryForm()
    return render(request, "advance_add.html", {"form": form})

@admin_required
def advance_delete(request, pk):
    adv = get_object_or_404(AdvanceSalary, pk=pk)
    if request.method == "POST":
        adv.delete()
        messages.success(request, "Advance removed.")
        return redirect("advance_list")
    return render(request, "advance_confirm_delete.html", {"advance": adv})

@login_required
def logout_view(request):
    logout(request)
    return redirect('login')

@login_required
def add_record(request):
    user = request.user
    is_admin = user.is_superuser

    manager_store = None
    if not is_admin and hasattr(user, "managerprofile"):
        manager_store = user.managerprofile.store

    if not is_admin and not manager_store:
        return HttpResponseForbidden("You are not allowed to add records.")

    if request.method == "POST":

        form = RecordForm(request.POST, is_admin=is_admin)

        if not form.is_valid():
            messages.error(request, f"Form errors: {form.errors}")

        else:
            date = form.cleaned_data.get("date")
            location = (
                form.cleaned_data.get("location")
                if is_admin
                else manager_store
            )

            vendors = request.POST.getlist("vendor[]")
            items = request.POST.getlist("item[]")
            quantities = request.POST.getlist("quantity[]")
            prices = request.POST.getlist("price[]")
            
            records_created = []

            try:
                with transaction.atomic():

                    for idx in range(len(vendors)):

                        v = vendors[idx]
                        i = items[idx] if idx < len(items) else None
                        q = quantities[idx] if idx < len(quantities) else None
                        p = prices[idx] if idx < len(prices) else None

                        if not (v and i and q and p):
                            continue

                        try:
                            quantity = Decimal(q)
                            unit_price = Decimal(p)
                        except Exception as e:
                            raise

                        try:
                            record = Record.objects.create(
                                date=date,
                                location=location,
                                vendor_id=v,
                                item_id=i,     # ← stays numeric only
                                quantity=quantity,
                                status="Pending",
                            )
                            records_created.append(record)
                        except Exception as e:
                            raise

                    if not records_created:
                        raise ValueError("No valid rows were saved")

                    for admin in User.objects.filter(is_superuser=True):
                        Notification.objects.create(
                            recipient=admin,
                            message=f"{len(records_created)} new inventory item(s) added",
                            location=location,
                            url=reverse("show_all_records"),
                        )

                messages.success(request, "Record saved successfully.")
                return redirect("add_record")

            except Exception as e:
                messages.error(request, f"Save failed: {e}")

    else:
        form = RecordForm(is_admin=is_admin)

    vendors = Vendor.objects.prefetch_related("items")

    return render(
        request,
        "addRecord.html",
        {
            "form": form,
            "vendors": vendors,
            "is_admin": is_admin,
            "manager_store": manager_store,
        }
    )

@login_required
def edit_order(request, pk):
    record = get_object_or_404(Record, pk=pk)

    # 🔒 Only admins/staff can edit orders
    if not request.user.is_staff:
        return redirect("record_detail", pk=pk)

    if request.method == "POST":
        record.quantity = int(request.POST["quantity"])
        record.item.unit_price = Decimal(request.POST["unit_price"])

        record.item.save()
        record.save()

        # 🔔 Notify managers of THIS STORE
        managers = User.objects.filter(
            managerprofile__store=record.location
        )

        for manager in managers:
            Notification.objects.create(
                recipient=manager,
                message=f"Order #{record.pk} was updated",
                location=record.location,  # Store
                url=reverse("record_detail", kwargs={"pk": record.pk}),
            )

    return redirect("record_detail", pk=pk)

@login_required
def fetch_notifications(request):
    user = request.user

    qs = (
        Notification.objects
        .filter(recipient=user)
        .order_by("-created_at")
    )

    data = []
    unread_count = 0

    for n in qs[:20]:
        if not n.is_read:
            unread_count += 1

        # 🔥 ADMIN → location filtered list
        if request.user.is_superuser:
            target_url = (
                reverse("show_all_records")
                + f"?location={n.location}"
                if n.location else reverse("show_all_records")
            )
        else:
            # 👤 MANAGER → order-level page
            target_url = n.url

        data.append({
            "id": n.id,
            "message": n.message,
            "location": n.location or "Unknown",
            "url": target_url,
            "has_unread": not n.is_read,
            "latest_time": n.created_at.strftime("%d %b, %H:%M"),
        })

    return JsonResponse({
        "notifications": data,
        "unread_count": unread_count,
        "is_admin": request.user.is_superuser
    })

@login_required
def clear_notifications(request):
    request.user.notifications.all().delete()
    return JsonResponse({"status": "ok"})

@login_required
def mark_notification_read(request, pk):
    notification = get_object_or_404(
        Notification,
        pk=pk,
        recipient=request.user
    )
    notification.is_read = True
    notification.save(update_fields=["is_read"])
    return JsonResponse({"ok": True})

def inventory_overview(request):
    qs = (
        Record.objects
        .filter(
            status="Completed",
            item__isnull=False
        )
        .values(
            "location",
            "item__item_name"
        )
        .annotate(total_quantity=Sum("quantity"))
        .order_by("location", "item__item_name")
    )

    inventory_by_store = defaultdict(list)

    for row in qs:
        inventory_by_store[row["location"]].append(row)

    return render(request, "inventory_overview.html", {
        "inventory_by_store": dict(inventory_by_store)
    })

@admin.register(Store)
class StoreAdmin(admin.ModelAdmin):
    list_display = ("name", "is_active")
    list_editable = ("is_active",)
    search_fields = ("name",)

def is_admin(user):
    return user.is_superuser

@user_passes_test(is_admin)
def manage_stores(request):
    if request.method == "POST":
        form = StoreWithManagerForm(request.POST)
        if form.is_valid():
            with transaction.atomic():

                # ------------------------
                # 1️⃣ Create Store
                # ------------------------
                store = Store.objects.create(
                    name=form.cleaned_data["store_name"],
                    is_active=form.cleaned_data.get("is_active", True)
                )

                # ------------------------
                # 2️⃣ Get or Create User
                # ------------------------
                username = form.cleaned_data["manager_username"]
                password = form.cleaned_data["manager_password"]

                user, created = User.objects.get_or_create(
                    username=username,
                    defaults={"email": ""}
                )

                # Set password only if user is newly created
                if created:
                    user.set_password(password)
                    user.save()

                # ------------------------
                # 3️⃣ Get or Create ManagerProfile
                # ------------------------
                profile, _ = ManagerProfile.objects.get_or_create(user=user)
                profile.store = store
                profile.is_active = True
                profile.save()

            return redirect("manage_stores")
    else:
        form = StoreWithManagerForm()

    stores = Store.objects.all().order_by("name")

    return render(request, "manage_stores.html", {
        "form": form,
        "stores": stores
    })

class StoreWithManagerForm(forms.Form):
    # ---- Store fields ----
    store_name = forms.CharField(max_length=100)
    is_active = forms.BooleanField(required=False, initial=True)

    # ---- Manager login fields ----
    manager_username = forms.CharField(max_length=150)
    manager_password = forms.CharField(widget=forms.PasswordInput)

    def clean_manager_username(self):
        username = self.cleaned_data["manager_username"]
        if User.objects.filter(username=username).exists():
            raise forms.ValidationError("Username already exists")
        return username